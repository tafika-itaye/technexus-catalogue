#!/usr/bin/env python3
"""Generate the <div class="content">...</div> block of catalogue.html
from the curated Catalogue + Icon_Search_Terms sheets in stock-selection-pricing.xlsx.

Usage:
  python3 scripts/build_catalogue.py           # EN  -> scripts/.catalogue_fragment.html
  python3 scripts/build_catalogue.py --lang pt # PT  -> scripts/.catalogue_fragment.pt.html
  python3 scripts/build_catalogue.py --lang ny # NY  -> scripts/.catalogue_fragment.ny.html
"""
from __future__ import annotations
import re
import sys
from pathlib import Path
from urllib.parse import quote
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "stock-selection-pricing.xlsx"

# Parse --lang flag
LANG = "en"
for _i, _arg in enumerate(sys.argv):
    if _arg == "--lang" and _i + 1 < len(sys.argv):
        LANG = sys.argv[_i + 1].lower()

OUT = REPO / "scripts" / (".catalogue_fragment.html" if LANG == "en"
                          else f".catalogue_fragment.{LANG}.html")

# For PT/NY, image paths must climb one level (../images/...)
PATH_PREFIX = "" if LANG == "en" else "../"

# UI strings per language
STRINGS = {
    "en": {
        "get_quote": "Get Quote",
        "wa_greeting": "Hi TechNexus, I would like a quote for:",
        "wa_price_label": "Price:",
        "ordering_h": "📋 Ordering Information",
        "ordering_p": ("All products are sourced through certified distributors. "
                       "Prices include Dubai→Malawi air freight and are valid for 30 days; "
                       "VAT (17.5%) is additional unless zero-rated. Volume discounts available "
                       "for schools, government, NGOs and enterprise. Contact TechNexus for "
                       "formal quotations, specifications and lead times."),
    },
    "pt": {
        "get_quote": "Pedir Orçamento",
        "wa_greeting": "Olá TechNexus, gostaria de um orçamento para:",
        "wa_price_label": "Preço:",
        "ordering_h": "📋 Informação de Encomenda",
        "ordering_p": ("Todos os produtos são adquiridos através de distribuidores certificados. "
                       "Os preços incluem frete aéreo Dubai→Malawi e são válidos durante 30 dias; "
                       "o IVA (17,5%) é adicional, salvo isenção. Descontos por volume disponíveis "
                       "para escolas, governo, ONGs e empresas. Contacte a TechNexus para "
                       "orçamentos formais, especificações e prazos de entrega."),
    },
    "ny": {
        "get_quote": "Pemphani Quotation",
        "wa_greeting": "Moni TechNexus, ndikufuna quotation ya:",
        "wa_price_label": "Mtengo:",
        "ordering_h": "📋 Zambiri za Kugula",
        "ordering_p": ("Zinthu zonse zimachokera kwa ogulitsa otsimikizika. "
                       "Mitengo ikuphatikiza ndege ya katundu kuchokera Dubai→Malawi ndipo "
                       "ndi yovomerezeka masiku 30; VAT (17.5%) imawonjezedwa pokhapokha "
                       "ngati pakhala zero-rated. Kuchotsera kwa katundu wambiri kumapezeka "
                       "ku masukulu, boma, ma NGO ndi mabizinesi. Lumikizanani ndi TechNexus "
                       "kuti mupeze ma quotation ovomerezeka, tsatanetsatane ndi nthawi zoperekera."),
    },
}
S = STRINGS[LANG]

# Section grouping (order on page) — (title, emoji, tagline, filter_fn)
# Curated title/specs overrides keyed by SKU (for items the auto-parser can't split cleanly)
OVERRIDES = {
    "MT-CCR2116-12G-4S+":   ("Mikrotik CCR2116-12G-4S+ Cloud Core Router",   "16GB RAM · 13x GbE · 4x SFP+ · RouterOS L6"),
    "MT-CCR2216-1G-12XS-2XQ": ("Mikrotik CCR2216 Cloud Core Router",         "12x SFP28 · 2x QSFP28 · 100G throughput · L6"),
    "MT-CCR2004-16G-2S+":   ("Mikrotik CCR2004-16G-2S+ Cloud Core Router",   "16x GbE · 2x SFP+ · RouterOS L5"),
    "MT-CCR2004-1G-12S+2XS": ("Mikrotik CCR2004-12S+2XS Aggregation Router", "12x SFP+ · 2x SFP28 · 25G uplink · L6"),
    "MT-RB5009UPR+S+IN":    ("Mikrotik RB5009UPR+S+IN PoE Router",           "7x GbE · 2.5G · SFP+ · PoE-in/out · desktop"),
    "MT-RB5009UG+S+IN":     ("Mikrotik RB5009UG+S+IN Router",                "7x GbE · 2.5G · SFP+ · desktop"),
    "MT-RB4011iGS+RM":      ("Mikrotik RB4011iGS+RM Rackmount Router",       "10x GbE · 1x SFP+ · 1U rackmount"),
    "MT-RB1100DX4":         ("Mikrotik RB1100AHx4 Dude Edition",             "13x GbE · 1U rackmount · Dude server preinstalled"),
    "MT-L009UIGS-RM":       ("Mikrotik L009UiGS-RM Rackmount Router",        "8x GbE · 2.5G · SFP · 1U rackmount"),
    "MT-CRS354-48G-4S+2Q+RM": ("Mikrotik CRS354-48G Managed Switch",         "48x GbE · 4x SFP+ · 2x QSFP+ · Layer 3 · Rack"),
    "MT-CRS354-48P-4S+2Q+RM": ("Mikrotik CRS354-48P PoE Switch",             "48x GbE PoE · 4x SFP+ · 2x QSFP+ · Layer 3 · Rack"),
    "MT-CRS328-24P-4S+RM":  ("Mikrotik CRS328-24P PoE Switch",               "24x GbE PoE+ · 4x SFP+ · Managed · Rack"),
    "MT-CRS326-24G-2S+RM":  ("Mikrotik CRS326-24G Managed Switch",           "24x GbE · 2x SFP+ · Layer 2/3 · Rack"),
    "MT-CRS326-24S+2Q+RM":  ("Mikrotik CRS326-24S+2Q+ Aggregation Switch",   "24x SFP+ 10G · 2x QSFP+ 40G · L3 · Rack"),
    "MT-CRS318-16P-2S+OUT": ("Mikrotik CRS318-16P NetPower Outdoor Switch",  "16x GbE PoE · 2x SFP+ · IP55 outdoor"),
    "MT-CRS112-8P-4S-IN":   ("Mikrotik CRS112-8P-4S Smart Switch",           "8x GbE PoE · 4x SFP · desktop"),
    "MT-S+85DLC03D":        ("Mikrotik S+85DLC03D SFP+ Transceiver",         "10GBASE-SR · 850nm multi-mode · 300m reach · LC"),
    "MT-S+RJ10":            ("Mikrotik S+RJ10 SFP+ Copper Module",           "10G RJ45 · Cat6a · 30m reach"),
}

# Per-section (key, emoji, filter_fn, icon_svg). Titles/taglines looked up by key+lang.
SECTIONS = [
    ("routers",      "📡", lambda sku, sub: sku.startswith("MT-") and "Router" in (sub or ""), "images/icons/router.svg"),
    ("switches",     "🔀", lambda sku, sub: sku.startswith("MT-CRS"),                          "images/icons/switch.svg"),
    ("transceivers", "🔌", lambda sku, sub: sku.startswith("MT-S+"),                           "images/icons/transceiver.svg"),
    ("laptops",      "💻", lambda sku, sub: sku.startswith(("LEN-TP-", "HP-PB")),              "images/icons/laptop.svg"),
    ("tower",        "🖥️", lambda sku, sub: sub == "Tower Desktop",                           "images/icons/desktop-tower.svg"),
    ("sff",          "🖥️", lambda sku, sub: sub == "SFF Desktop",                             "images/icons/desktop-sff.svg"),
    ("mini",         "📦", lambda sku, sub: sub == "Mini PC",                                  "images/icons/mini-pc.svg"),
]

SECTION_TEXT = {
    "routers": {
        "en": ("Enterprise & ISP Routers",
               "Mikrotik Cloud Core and RouterBOARD for ISPs, data centres and enterprise edge."),
        "pt": ("Routers Empresariais e para ISP",
               "Mikrotik Cloud Core e RouterBOARD para ISPs, centros de dados e redes empresariais."),
        "ny": ("Ma Router a Bizinesi ndi a ISP",
               "Mikrotik Cloud Core ndi RouterBOARD za ma ISP, ma data centre ndi ma netiweki a bizinesi."),
    },
    "switches": {
        "en": ("Managed Switches",
               "Layer-2/3 managed switches — gigabit, 10G/40G aggregation and PoE distribution."),
        "pt": ("Switches Geridos",
               "Switches geridos Layer-2/3 — gigabit, agregação 10G/40G e distribuição PoE."),
        "ny": ("Ma Switch Oyendetsedwa",
               "Ma switch oyendetsedwa a Layer-2/3 — gigabit, kusakaniza kwa 10G/40G ndi kugawa kwa PoE."),
    },
    "transceivers": {
        "en": ("SFP+ Transceivers",
               "10G fibre and copper modules for your Mikrotik, HP, Dell or Cisco switches."),
        "pt": ("Transceivers SFP+",
               "Módulos 10G de fibra e cobre para os seus switches Mikrotik, HP, Dell ou Cisco."),
        "ny": ("Ma Transceiver a SFP+",
               "Ma module a 10G a faibala ndi a mkuwa a ma switch anu a Mikrotik, HP, Dell kapena Cisco."),
    },
    "laptops": {
        "en": ("Business Laptops",
               "Lenovo ThinkPad and HP ProBook — warranty-backed business-class notebooks."),
        "pt": ("Laptops Empresariais",
               "Lenovo ThinkPad e HP ProBook — notebooks de classe empresarial com garantia."),
        "ny": ("Malaputopu a Bizinesi",
               "Lenovo ThinkPad ndi HP ProBook — ma notebook a bizinesi okhala ndi chitsimikizo cha warranty."),
    },
    "tower": {
        "en": ("Business Desktops — Tower",
               "Full-size towers for offices, labs and control rooms."),
        "pt": ("Desktops Empresariais — Torre",
               "Torres de tamanho grande para escritórios, laboratórios e salas de controlo."),
        "ny": ("Ma Desktop a Bizinesi — Tower",
               "Ma tower akuluakulu a ma ofesi, ma labu ndi zipinda zolamulira."),
    },
    "sff": {
        "en": ("Business Desktops — Small Form Factor",
               "Space-efficient SFF desktops for dense office deployments."),
        "pt": ("Desktops Empresariais — Formato Compacto (SFF)",
               "Desktops SFF que poupam espaço para escritórios densamente equipados."),
        "ny": ("Ma Desktop a Bizinesi — Ocheperako (SFF)",
               "Ma desktop a SFF osunga malo a ma ofesi okhala ndi makompyuta ambiri."),
    },
    "mini": {
        "en": ("Business Desktops — Mini PC",
               "Compact Mini/Tiny PCs for signage, reception and hot-desk setups."),
        "pt": ("Desktops Empresariais — Mini PC",
               "Mini PCs compactos para sinalização digital, recepção e secretárias partilhadas."),
        "ny": ("Ma Desktop a Bizinesi — Mini PC",
               "Ma Mini PC ocheperako a zizindikiro za digito, malo olandirira alendo ndi ma desk ogawana."),
    },
}

# Split a product name into (short_title, spec_line).
# Rules:
#  1. Strip trailing vendor SKU parentheticals like (21T9006C1G), (CA7Y3AT#BH5)
#  2. If a remaining parenthetical exists, treat it as the spec line
#  3. Else split on first CPU/RAM marker (Core, Ultra, Ryzen, AMD, 16GB, etc.)
CPU_RE = re.compile(
    r"\s+(Core [0-9iI]|Ultra [0-9]|AMD\b|Ryzen|i[3579]-|R[5-9]-|Intel Core|Celeron)",
    re.IGNORECASE,
)
SIZE_RE = re.compile(r"\s+(\d+GB/|\d+\s*GB\s)")

def split_name(name: str) -> tuple[str, str]:
    # 1. Strip trailing vendor SKU in parens
    name = re.sub(r"\s*\([A-Z0-9#]{4,}(?:\s*[A-Z0-9#]+)*\)\s*$", "", name).strip()
    # 2. Remaining parenthetical → specs
    m = re.search(r"^(.*?)\s*\((.+?)\)\s*$", name)
    if m:
        title = m.group(1).strip()
        specs = m.group(2).strip()
    else:
        # 3. Try CPU marker split
        m = CPU_RE.search(name)
        if not m:
            m = SIZE_RE.search(name)
        if m:
            title = name[:m.start()].strip()
            specs = name[m.start():].strip()
        else:
            title = name
            specs = ""
    specs = re.sub(r",\s*", " · ", specs)
    specs = re.sub(r"\s+/\s+", "/", specs)
    # Put bullets between words that should be separated: 14" DOS, 512GB DOS, etc.
    specs = re.sub(r"(\d+GB/\d+GB)\s+(\d+\")", r"\1 · \2", specs)
    specs = re.sub(r"(\d+\")\s+(DOS|Win\b|FHD|WUXGA|IPS)", r"\1 · \2", specs)
    specs = re.sub(r"\s+(DOS|Win 11|Win 10|FHD|BL|FPR)\b", r" · \1", specs)
    specs = re.sub(r"·\s+·", "·", specs)
    return title, specs

def card_html(sku_row, icon_row, icon_svg):
    (sku, cat, sub, name, brand, supplier,
     cost, src, weight, freight, landed, final_usd, final_mwk, include) = sku_row
    # Primary image strategy:
    #   - If a real vendor photo exists at the per-SKU path on disk, USE IT.
    #   - Else fall back to the category SVG placeholder (still visible, professional).
    filename_base = icon_row[1] if icon_row else sku.lower().replace("+","-").replace("_","-")
    folder = icon_row[2] if icon_row else "images/misc/"
    per_sku_path = f"{folder}{filename_base}.jpg"
    per_sku_png  = f"{folder}{filename_base}.png"
    if (REPO / per_sku_path).exists():
        img_path = per_sku_path
    elif (REPO / per_sku_png).exists():
        img_path = per_sku_png
    else:
        img_path = icon_svg
    # Prefix for PT/NY (files live one level deep)
    img_src = PATH_PREFIX + img_path
    data_sku_img = PATH_PREFIX + per_sku_path
    if sku in OVERRIDES:
        title, specs = OVERRIDES[sku]
    else:
        title, specs = split_name(name)
    usd = f"${int(final_usd):,}"
    mwk = f"MK {int(final_mwk):,}"
    wa_text = f"{S['wa_greeting']}\n{title}\nSKU: {sku}\n{S['wa_price_label']} {usd}"
    wa_href = "https://wa.me/265889941700?text=" + quote(wa_text, safe="")
    return (
        f'<div class="card">'
        f'<div class="card-img-wrap">'
        f'<img src="{img_src}" data-sku-img="{data_sku_img}" alt="{title}" loading="lazy">'
        f'</div>'
        f'<div class="card-body">'
        f'<div class="sku">{sku}</div>'
        f'<div class="card-title">{title}</div>'
        f'<div class="specs">{specs}</div>'
        f'<span class="price-mwk">{mwk}</span>'
        f'<span class="price-usd">{usd}</span>'
        f'<a class="wa-quote" href="{wa_href}" target="_blank" rel="noopener">'
        f'<svg width="16" height="16" aria-hidden="true"><use href="#wa-icon" fill="currentColor"/></svg>'
        f' {S["get_quote"]}</a>'
        f'</div></div>'
    )

def main():
    wb = load_workbook(XLSX, data_only=True)
    cat = wb["Catalogue"]
    ic  = wb["Icon_Search_Terms"]

    # Build icon lookup { SKU: row }
    icons = {}
    for row in ic.iter_rows(min_row=4, values_only=True):  # header at row 3
        if row and row[0]:
            icons[row[0]] = row

    # Collect catalogue rows with Include == 'Yes'
    rows = []
    for row in cat.iter_rows(min_row=2, values_only=True):
        if row and row[0] and (row[13] or "").strip().lower() == "yes":
            rows.append(row)

    out = ['<div class="content">\n']
    for key, emoji, flt, icon_svg in SECTIONS:
        items = [r for r in rows if flt(r[0], r[2])]
        if not items:
            continue
        title, tagline = SECTION_TEXT[key][LANG]
        out.append(f'\n  <div class="sh"><h2>{emoji} {title}</h2><p class="sd">{tagline}</p></div>\n')
        out.append('  <div class="card-grid">\n')
        for r in items:
            out.append("    " + card_html(r, icons.get(r[0]), icon_svg) + "\n")
        out.append('  </div>\n')

    out.append(f'\n  <div class="compliance"><h3>{S["ordering_h"]}</h3><p>{S["ordering_p"]}</p></div>\n')
    out.append('</div>\n')

    OUT.write_text("".join(out), encoding="utf-8")
    total = sum(1 for line in out if '<div class="card">' in line)
    print(f"[{LANG}] Wrote {OUT.relative_to(REPO)} — {total} cards across {sum(1 for _k, _e, f, _s in SECTIONS if any(f(r[0], r[2]) for r in rows))} sections")

if __name__ == "__main__":
    main()
